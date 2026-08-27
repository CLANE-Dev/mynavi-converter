"""パスワード付き xlsx を復号し、自社情報と角印を入れて保存する。"""
from __future__ import annotations

import io
import logging
import os

import msoffcrypto
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import coordinate_to_tuple
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PILImage

import config

log = logging.getLogger(__name__)


class TemplateChanged(Exception):
    """テンプレートの様式が想定と違う。変換せず止めるための例外。"""


def decrypt_to_bytes(raw: bytes, password: str) -> bytes:
    """暗号化 xlsx を復号して平文 xlsx のバイト列を返す。"""
    src = io.BytesIO(raw)
    try:
        office = msoffcrypto.OfficeFile(src)
    except Exception:
        src.seek(0)
        return raw

    if not office.is_encrypted():
        return raw

    office.load_key(password=password)
    out = io.BytesIO()
    office.decrypt(out)
    return out.getvalue()


def load(raw: bytes, password: str) -> openpyxl.Workbook:
    plain = decrypt_to_bytes(raw, password)
    return openpyxl.load_workbook(io.BytesIO(plain))


def _target_sheet(wb: openpyxl.Workbook):
    if config.TARGET_SHEET in wb.sheetnames:
        return wb[config.TARGET_SHEET]
    log.warning("シート %s が見つからないため先頭シートを使用", config.TARGET_SHEET)
    return wb.worksheets[0]


def check_template(wb: openpyxl.Workbook, is_invoice: bool) -> None:
    """想定ラベルが揃っているか確認する。欠けていれば TemplateChanged。"""
    ws = _target_sheet(wb)
    expected = dict(config.EXPECTED_LABELS)
    if is_invoice:
        expected.update(config.EXPECTED_LABELS_INVOICE)

    missing = []
    for addr, keywords in expected.items():
        cur = ws[addr].value
        text = "" if cur is None else str(cur)
        if not any(k in text for k in keywords):
            missing.append(f"{addr}(期待:{'/'.join(keywords)} 実際:{text.strip()[:20] or '空'})")

    if missing:
        raise TemplateChanged("想定と異なるセル: " + " , ".join(missing))


def _write(ws, addr: str, value: str) -> None:
    """既存ラベルを保持して書き込む。"""
    if value is None or value == "":
        return
    cur = ws[addr].value
    if cur is None or cur == "":
        ws[addr] = value
        return
    cur = str(cur)
    if cur.rstrip().endswith(("：", ":")):
        ws[addr] = cur + value
        return
    if value in cur:
        return
    if cur.strip() == "㊞":
        return
    if cur.rstrip().endswith("〒"):
        ws[addr] = cur.rstrip() + value.lstrip("〒")
        return
    ws[addr] = cur + value


def _add_stamp(ws, stamp_png: bytes) -> None:
    """角印を表紙の押印欄に貼る。縦横比は維持し、セル基準＋画素オフセットで置く。"""
    with PILImage.open(io.BytesIO(stamp_png)) as im:
        w, h = im.size
    width = config.STAMP_WIDTH_PX
    height = int(width * h / w)

    img = XLImage(io.BytesIO(stamp_png))
    img.width = width
    img.height = height

    row, col = coordinate_to_tuple(config.STAMP_ANCHOR)
    marker = AnchorMarker(
        col=col - 1,
        colOff=pixels_to_EMU(config.STAMP_OFFSET_X_PX),
        row=row - 1,
        rowOff=pixels_to_EMU(config.STAMP_OFFSET_Y_PX),
    )
    img.anchor = OneCellAnchor(
        _from=marker,
        ext=XDRPositiveSize2D(pixels_to_EMU(width), pixels_to_EMU(height)),
    )
    ws.add_image(img)


def _drop_excluded_sheets(wb) -> list:
    """注意書き等、送付対象外のシートをPDF化前に取り除く。"""
    removed = []
    for name in list(wb.sheetnames):
        if name in config.EXCLUDE_SHEETS and len(wb.sheetnames) > 1:
            wb.remove(wb[name])
            removed.append(name)
    return removed


def _fit_to_one_page(wb) -> None:
    """全シートを横1×縦1ページに収める。"""
    for ws in wb.worksheets:
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_setup.scale = None
        ws.sheet_properties.pageSetUpPr.fitToPage = True


def edit(
    raw: bytes,
    *,
    password: str,
    is_invoice: bool,
    stamp_png: bytes | None,
    out_path: str,
) -> dict:
    """記入済み xlsx を out_path に書き出し、書き込んだ内容を返す。"""
    wb = load(raw, password)
    check_template(wb, is_invoice)

    ws = _target_sheet(wb)
    cells = dict(config.COMMON_CELLS)
    if is_invoice:
        cells.update(config.INVOICE_CELLS)

    written = {}
    for addr, value in cells.items():
        _write(ws, addr, value)
        written[addr] = ws[addr].value

    stamped = False
    if stamp_png:
        try:
            _add_stamp(ws, stamp_png)
            stamped = True
        except Exception as exc:
            log.exception("角印の挿入に失敗: %s", exc)

    removed = _drop_excluded_sheets(wb)
    if config.FIT_TO_PAGE:
        _fit_to_one_page(wb)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return {
        "sheet": ws.title,
        "written": written,
        "stamped": stamped,
        "removed_sheets": removed,
        "sheets": wb.sheetnames,
    }
